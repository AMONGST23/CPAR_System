from django.contrib.auth import get_user_model
from django.contrib.auth.models import User
from django.urls import reverse
from django.test import TestCase
from openpyxl import load_workbook
from io import BytesIO

from surveys.models import MaternalRecord
from .models import AgentProfile


class AgentProfileSignalTests(TestCase):
    def test_profile_created_with_user(self):
        user = User.objects.create_user(username='agent1', password='secret123')
        self.assertTrue(AgentProfile.objects.filter(user=user).exists())


class AdminPanelTests(TestCase):
    def setUp(self):
        self.staff_user = get_user_model().objects.create_user(
            username='admin1',
            password='secret123',
            is_staff=True,
        )
        self.agent_user = get_user_model().objects.create_user(
            username='agent1',
            password='secret123',
        )

    def test_staff_user_can_open_admin_panel(self):
        self.client.login(username='admin1', password='secret123')
        response = self.client.get(reverse('admin_panel'))
        self.assertEqual(response.status_code, 200)

    def test_staff_user_can_create_agent(self):
        self.client.login(username='admin1', password='secret123')
        response = self.client.post(reverse('admin_panel'), {
            'action': 'create_agent',
            'create-username': 'agent2',
            'create-first_name': 'Jane',
            'create-last_name': 'Doe',
            'create-assigned_area': 'Ward 1',
            'create-phone_number': '0700000000',
            'create-password1': 'secret12345',
            'create-password2': 'secret12345',
        })
        self.assertEqual(response.status_code, 302)
        self.assertTrue(get_user_model().objects.filter(username='agent2').exists())

    def test_admin_can_export_records_to_excel(self):
        MaternalRecord.objects.create(
            last_name='Doe',
            first_name='Jane',
            date_collected='2026-03-24',
            address_barangay='Nairobi',
            is_currently_pregnant=True,
            ultrasound_before_24_weeks=True,
            agent=self.agent_user,
        )

        self.client.login(username='admin1', password='secret123')
        response = self.client.get(reverse('admin_export_records'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook.active
        headers = [cell.value for cell in worksheet[1]]
        self.assertIn('Region', headers)
        self.assertIn('Marital Status', headers)
        self.assertIn('Occupation', headers)
        self.assertIn('Ultrasound Before 24 Weeks', headers)
        self.assertNotIn('Educational Attainment', headers)
        self.assertNotIn('High-Risk Identified Through RUAA', headers)

        first_row = [cell.value for cell in worksheet[2]]
        self.assertIn('Doe', first_row)
        self.assertIn('Nairobi', first_row)
        self.assertIn('Yes', first_row)

    def test_non_admin_cannot_export_records_to_excel(self):
        self.client.login(username='agent1', password='secret123')
        response = self.client.get(reverse('admin_export_records'))
        self.assertEqual(response.status_code, 302)

    def test_admin_can_filter_records_by_agent(self):
        second_agent = get_user_model().objects.create_user(
            username='agent2',
            password='secret123',
            first_name='Second',
            last_name='Agent',
        )
        MaternalRecord.objects.create(
            last_name='Doe',
            first_name='Jane',
            date_collected='2026-03-24',
            agent=self.agent_user,
        )
        MaternalRecord.objects.create(
            last_name='Smith',
            first_name='John',
            date_collected='2026-03-24',
            agent=second_agent,
        )

        self.client.login(username='admin1', password='secret123')
        response = self.client.get(reverse('admin_panel'), {'agent': str(self.agent_user.id)})

        self.assertEqual(response.status_code, 200)
        records = list(response.context['records'])
        self.assertEqual(len(records), 1)
        self.assertEqual(records[0].agent, self.agent_user)

    def test_excel_export_respects_agent_filter(self):
        second_agent = get_user_model().objects.create_user(
            username='agent2',
            password='secret123',
        )
        MaternalRecord.objects.create(
            last_name='Doe',
            first_name='Jane',
            date_collected='2026-03-24',
            agent=self.agent_user,
        )
        MaternalRecord.objects.create(
            last_name='Smith',
            first_name='John',
            date_collected='2026-03-24',
            agent=second_agent,
        )

        self.client.login(username='admin1', password='secret123')
        response = self.client.get(reverse('admin_export_records'), {'agent': str(self.agent_user.id)})

        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))

        self.assertEqual(len(rows), 2)
        self.assertIn('Doe', rows[1])
        self.assertNotIn('Smith', rows[1])

    def test_admin_panel_shows_agent_record_summary(self):
        second_agent = get_user_model().objects.create_user(
            username='agent2',
            password='secret123',
        )
        MaternalRecord.objects.create(
            last_name='Doe',
            first_name='Jane',
            date_collected='2026-03-24',
            agent=self.agent_user,
        )
        MaternalRecord.objects.create(
            last_name='Smith',
            first_name='John',
            date_collected='2026-03-24',
            agent=self.agent_user,
        )
        MaternalRecord.objects.create(
            last_name='Brown',
            first_name='Anne',
            date_collected='2026-03-24',
            agent=second_agent,
        )

        self.client.login(username='admin1', password='secret123')
        response = self.client.get(reverse('admin_panel'))

        self.assertEqual(response.status_code, 200)
        summaries = response.context['agent_summaries']
        summary_by_username = {item['user'].username: item['record_count'] for item in summaries}
        self.assertEqual(summary_by_username['agent1'], 2)
        self.assertEqual(summary_by_username['agent2'], 1)
